#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
data/ai-research-feed.json を読み、参考スプレッドシート
https://docs.google.com/spreadsheets/d/1pMTIWgWfPFEUkOh4V7KkDsSSYEeSlBDVNr9_zFxBm_o/edit?gid=905697318
と同じ列構成(日付, カテゴリ, タイトル, 概要, ソースURL, 収集日時, 使うAI)のExcelを生成する。

使い方:
    python3 scripts/build_ai_research_feed_xlsx.py [出力パス]
    (出力パス省略時は ./ai-research-feed.xlsx)

「AI技術リサーチ Routine」(毎日JST9:00、このリポジトリの相談ルームへ自己バインド)が
日々このスクリプトを実行し、生成したxlsxをユーザーへ送付する運用を想定している。
"""
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(HERE, "..", "data", "ai-research-feed.json")

HEADERS = ["日付", "カテゴリ", "タイトル", "概要", "ソースURL", "収集日時", "使うAI"]
COL_WIDTHS = [12, 20, 45, 55, 45, 16, 10]


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else "ai-research-feed.xlsx"

    with open(DATA_PATH, encoding="utf-8") as f:
        data = json.load(f)
    entries = data.get("entries", [])
    # 日付の新しい順、同日内は追加順を保つ
    entries_sorted = sorted(entries, key=lambda e: e.get("date", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "AI技術リサーチ"

    ws.append(HEADERS)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    row_idx = 1
    for e in entries_sorted:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=e.get("date", ""))
        ws.cell(row=row_idx, column=2, value=e.get("category", ""))
        ws.cell(row=row_idx, column=3, value=e.get("title", ""))
        ws.cell(row=row_idx, column=4, value=e.get("summary", ""))
        ws.cell(row=row_idx, column=5, value=e.get("sourceUrl", ""))
        ws.cell(row=row_idx, column=6, value=e.get("collectedAt", ""))
        ws.cell(row=row_idx, column=7, value=e.get("usedAI", ""))
        for col in [3, 4]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # サマリーシート
    ws2 = wb.create_sheet("サマリー")
    ws2.append(["項目", "値"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    by_date = {}
    for e in entries:
        by_date.setdefault(e.get("date", "?"), 0)
        by_date[e.get("date", "?")] += 1
    ws2.append(["総件数", len(entries)])
    ws2.append(["収集日数", len(by_date)])
    for d in sorted(by_date.keys(), reverse=True)[:10]:
        ws2.append([d, by_date[d]])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16

    wb.save(out_path)
    print(f"saved {out_path} ({len(entries)} rows)")


if __name__ == "__main__":
    main()
